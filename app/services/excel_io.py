"""Импорт/экспорт журнала в формате общего Excel-журнала участка.

Формат листа журнала (Sheet1):
Дата | Время | Изделие | Акт | Кол-во | Акт несоот. | Кол-во рем. | Статус |
к Отгрузке (дата) | к Отгрузке (время) | факт Отгр (дата) | факт Отгр (время) |
Выполнил вых кон | Примечание
"""

import re
from datetime import date, datetime, time
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models import Batch, StatusEvent, SyncLog

HEADERS = [
    "Дата", "Время", "Изделие", "Акт", "Кол-во", "Акт несоот.", "Кол-во рем.",
    "Статус", "к Отгрузке (дата)", "к Отгрузке (время)", "факт Отгр (дата)",
    "факт Отгр (время)", "Выполнил вых кон", "Примечание",
]

# нормализация статусов из Excel (там встречаются лишние пробелы и вариации)
_CANON_STATUSES = {
    "принят": "Принят",
    "входной контроль": "Входной контроль",
    "в работе": "В работе",
    "остановлен": "Остановлен",
    "выходной контроль": "Выходной контроль",
    "к отгрузке": "К отгрузке",
    "отгружен": "Отгружен",
    "отгружено": "Отгружен",
}


def normalize_status(raw) -> str:
    if not raw:
        return "Принят"
    key = re.sub(r"\s+", " ", str(raw)).strip().lower()
    return _CANON_STATUSES.get(key, str(raw).strip())


def _to_date(val) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    # встречаются кривые записи вроде '20:07.26' → 20.07.(20)26
    m = re.match(r"^(\d{1,2})[.:/-](\d{1,2})[.:/-](\d{2,4})$", s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y < 100:
            y += 2000
        try:
            return date(y, mo, d)
        except ValueError:
            return None
    return None


def _to_time(val) -> time | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        # openpyxl отдаёт время >24ч как datetime от 1900 года
        return val.time()
    if isinstance(val, time):
        return val
    m = re.match(r"^(\d{1,2})[.:](\d{2})", str(val).strip())
    if m:
        try:
            return time(int(m.group(1)) % 24, int(m.group(2)))
        except ValueError:
            return None
    return None


def _combine(d, t) -> datetime | None:
    dd, tt = _to_date(d), _to_time(t)
    if dd is None and tt is None:
        return None
    return datetime.combine(dd or date.today(), tt or time(0, 0))


def _to_int(val) -> int:
    try:
        return int(float(val))
    except (TypeError, ValueError):
        return 0


def _find_journal_sheet(wb):
    """Находит лист журнала по заголовку 'Акт' + 'Изделие' в первой строке."""
    for ws in wb.worksheets:
        first = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
        if any("акт" == v for v in first) and any("изделие" in v for v in first):
            return ws
    return wb.worksheets[0]


def import_journal(db: Session, path: str | Path, source: str = "") -> SyncLog:
    """Импорт журнала. Ключ — (№ акта, изделие): существующие записи
    обновляются, новые создаются. Дубли не плодятся."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = _find_journal_sheet(wb)

    header_map = {}
    for idx, cell in enumerate(ws[1]):
        if cell.value:
            header_map[re.sub(r"\s+", " ", str(cell.value)).strip().lower()] = idx

    def col(row, *names):
        for n in names:
            if n in header_map and header_map[n] < len(row):
                return row[header_map[n]]
        return None

    created = updated = skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        act = col(row, "акт")
        product = col(row, "изделие")
        if act is None or product is None:
            skipped += 1
            continue
        act = str(act).strip() if not isinstance(act, float) else str(int(act))
        product = str(product).strip()
        if not act or not product:
            skipped += 1
            continue

        data = dict(
            quantity=_to_int(col(row, "кол-во")),
            status=normalize_status(col(row, "статус")),
            accepted_at=_combine(col(row, "дата"), col(row, "время")),
            planned_ship_at=_combine(
                col(row, "к отгрузке (дата)"), col(row, "к отгрузке (время)")
            ),
            actual_ship_at=_combine(
                col(row, "факт отгр (дата)"), col(row, "факт отгр (время)")
            ),
            nc_act_number=(
                str(col(row, "акт несоот.", "акт несоот")).strip()
                if col(row, "акт несоот.", "акт несоот") is not None
                else None
            ),
            repair_qty=_to_int(col(row, "кол-во рем.", "кол-во рем")),
            output_control_by=(
                str(col(row, "выполнил вых кон")).strip()
                if col(row, "выполнил вых кон") is not None
                else None
            ),
            note=(
                str(col(row, "примечание")).strip()
                if col(row, "примечание") is not None
                else None
            ),
        )

        batch = (
            db.query(Batch)
            .filter(Batch.act_number == act, Batch.product == product)
            .first()
        )
        if batch is None:
            batch = Batch(act_number=act, product=product, **data)
            db.add(batch)
            db.flush()
            db.add(
                StatusEvent(
                    batch_id=batch.id, status=batch.status, comment="Импорт из Excel"
                )
            )
            created += 1
        else:
            changed = False
            old_status = batch.status
            for k, v in data.items():
                if v not in (None, 0, "") and getattr(batch, k) != v:
                    setattr(batch, k, v)
                    changed = True
            if changed:
                if batch.status != old_status:
                    db.add(
                        StatusEvent(
                            batch_id=batch.id,
                            status=batch.status,
                            comment="Импорт из Excel",
                        )
                    )
                updated += 1
            else:
                skipped += 1

    log = SyncLog(
        direction="import",
        source=source or str(path),
        created=created,
        updated=updated,
        skipped=skipped,
    )
    db.add(log)
    db.commit()
    return log


# ---------------------------------------------------------------- export

_THIN = Border(*[Side(style="thin")] * 4)
_HDR_FILL = PatternFill("solid", start_color="D9E1F2")
_STATUS_FILLS = {
    "Отгружен": PatternFill("solid", start_color="C6EFCE"),
    "К отгрузке": PatternFill("solid", start_color="FFEB9C"),
    "Остановлен": PatternFill("solid", start_color="FFC7CE"),
}


def export_journal(db: Session, path: str | Path) -> SyncLog:
    """Экспорт: лист-журнал в общем формате + сводка по изделиям/актам."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Журнал"

    for i, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = _HDR_FILL
        c.border = _THIN
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    batches = (
        db.query(Batch).order_by(Batch.accepted_at.desc().nullslast(), Batch.id).all()
    )
    for r, b in enumerate(batches, 2):
        vals = [
            b.accepted_at.date() if b.accepted_at else None,
            b.accepted_at.time() if b.accepted_at else None,
            b.product,
            b.act_number,
            b.quantity,
            b.nc_act_number,
            b.repair_qty or None,
            b.status,
            b.planned_ship_at.date() if b.planned_ship_at else None,
            b.planned_ship_at.time() if b.planned_ship_at else None,
            b.actual_ship_at.date() if b.actual_ship_at else None,
            b.actual_ship_at.time() if b.actual_ship_at else None,
            b.output_control_by,
            b.note,
        ]
        for i, v in enumerate(vals, 1):
            c = ws.cell(row=r, column=i, value=v)
            c.border = _THIN
            if isinstance(v, date) and not isinstance(v, datetime):
                c.number_format = "DD.MM.YYYY"
            elif isinstance(v, time):
                c.number_format = "HH:MM"
        fill = _STATUS_FILLS.get(b.status)
        if fill:
            ws.cell(row=r, column=8).fill = fill

    widths = [11, 7, 14, 9, 8, 11, 10, 18, 14, 13, 13, 13, 14, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:N{max(len(batches) + 1, 2)}"

    # ---- сводка (аналог их сводной таблицы: изделие → акты, суммы)
    sv = wb.create_sheet("Сводка")
    sv.cell(row=1, column=1, value="Изделие / Акт").font = Font(bold=True)
    sv.cell(row=1, column=2, value="Кол-во").font = Font(bold=True)
    sv.cell(row=1, column=3, value="Кол-во рем.").font = Font(bold=True)
    for c in sv[1]:
        c.fill = _HDR_FILL
        c.border = _THIN

    r = 2
    total_q = total_r = 0
    products = (
        db.query(
            Batch.product,
            func.sum(Batch.quantity),
            func.sum(Batch.repair_qty),
        )
        .group_by(Batch.product)
        .order_by(Batch.product)
        .all()
    )
    for product, q, rq in products:
        c = sv.cell(row=r, column=1, value=product)
        c.font = Font(bold=True)
        sv.cell(row=r, column=2, value=q or 0).font = Font(bold=True)
        sv.cell(row=r, column=3, value=rq or 0).font = Font(bold=True)
        r += 1
        for b in (
            db.query(Batch)
            .filter(Batch.product == product)
            .order_by(Batch.act_number)
            .all()
        ):
            sv.cell(row=r, column=1, value="    " + str(b.act_number))
            sv.cell(row=r, column=2, value=b.quantity)
            sv.cell(row=r, column=3, value=b.repair_qty or None)
            r += 1
        total_q += q or 0
        total_r += rq or 0

    c = sv.cell(row=r, column=1, value="Общий итог")
    c.font = Font(bold=True)
    sv.cell(row=r, column=2, value=total_q).font = Font(bold=True)
    sv.cell(row=r, column=3, value=total_r).font = Font(bold=True)
    sv.column_dimensions["A"].width = 22
    sv.column_dimensions["B"].width = 10
    sv.column_dimensions["C"].width = 12

    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(".tmp.xlsx")
    wb.save(tmp)
    tmp.replace(path)

    log = SyncLog(direction="export", source=str(path), created=len(batches))
    db.add(log)
    db.commit()
    return log
